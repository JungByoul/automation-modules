# -*- coding: utf-8 -*-
"""
Outlook(로컬 COM) 받은편지함에서 infowise.kr 발신 mMDM FINAL REVIEW 알림메일(패턴 A)을
읽어 MOC 관리대장 xlsx로 저장한다.

- 인증/MCP 불필요. 로컬 클래식 Outlook COM 사용.
- 접수일(받은날짜) CUTOFF 이후 메일만 수집 (기본: 2026-07-01 이후 → 7월부터).
- 실행할 때마다 _메일수집_결과 폴더에 타임스탬프 새 파일 생성 (버전 스냅샷).
- (Workflow ID + 설비ID) 중복은 한 번 실행 내에서 스킵.
"""
import win32com.client as win32
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re, datetime, os, time

# ---------------------------------------------------------------- 경로/설정
_HOME = os.path.expanduser("~")
BASE = os.path.join(_HOME, r"1_Work\1.20_S-oil_Projects\1.20.20_mMDM_MOC_관리대장(아카이브)")
OUT_RESULT_DIR = os.path.join(BASE, "1.20.20.10_메일수집_결과")   # 실행 결과 저장 폴더
SENDER_KEY = "infowise"                                 # 발신자 SMTP/이름 필터
CUTOFF = datetime.date(2026, 7, 1)                      # 이 날짜 이후(포함) 접수건만
SYNC_WAIT = 8                                           # 실행 시 강제 동기화 후 대기(초). 방금 온 메일도 잡히게
ADDRBOOK = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "메일주소_주소록.xlsx")         # 메일→이름 매칭용 주소록 (없으면 매칭 없이 진행)

# 컬럼 순서: 설비ID 오른쪽에 등록요청관련·비고, 비고 오른쪽에 접수 시간
HEADER = ["no.", "접수일", "완료일(월)", "구분", "Workflow ID", "설비ID",
          "등록요청관련", "비고", "접수 시간", "Criticality", "기술식별번호",
          "SRCM DB 파일명", "RTS 대상 여부 및 확인", "MDM review 이후 확인", ""]

# ---------------------------------------------------------------- 파싱 함수
def smtp_of(item):
    try:
        return item.PropertyAccessor.GetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x5D01001F")
    except Exception:
        return getattr(item, "SenderEmailAddress", "") or ""

def parse_equipment_list(body):
    lines = body.splitlines()
    hdr_idx = None
    for i, ln in enumerate(lines):
        cells = [c.strip() for c in ln.split("\t")]
        if "설비ID" in cells and "Workflow ID" in cells:
            hdr_idx = i
            break
    if hdr_idx is None:
        return []
    h = [c.strip() for c in lines[hdr_idx].split("\t") if c.strip()]
    rows = []
    for ln in lines[hdr_idx + 1:]:
        if not ln.strip():
            break
        cells = [c.strip() for c in ln.split("\t") if c.strip() != ""]
        if not cells:
            break
        rows.append(dict(zip(h, cells)))
    return rows

def load_addrbook():
    """주소록 xlsx 전체 시트에서 {메일주소(소문자): 이름} 사전 생성.
    각 시트 헤더 행에서 '메일' 포함 컬럼과 '요청자'/'담당자' 포함 컬럼을 찾아 매핑.
    파일이 없거나 읽기 실패 시 빈 dict → 기존 동작(메일 원문 표기)으로 폴백."""
    addr_map = {}
    try:
        wb = openpyxl.load_workbook(ADDRBOOK, read_only=True, data_only=True)
    except Exception as e:
        print(f"주소록 로드 실패({e}) → 메일 원문 사용")
        return addr_map
    for ws in wb.worksheets:
        mail_col = name_col = None
        for row in ws.iter_rows(values_only=True):
            if mail_col is None:                 # 첫 행에서 헤더 위치 탐색
                for i, v in enumerate(row):
                    h = str(v or "").strip()
                    if "메일" in h:
                        mail_col = i
                    elif "요청자" in h or "담당자" in h:
                        name_col = i
                if mail_col is None or name_col is None:
                    break                        # 헤더를 못 찾는 시트는 건너뜀
                continue
            mail = str(row[mail_col] or "").strip().lower() if len(row) > mail_col else ""
            name = str(row[name_col] or "").strip() if len(row) > name_col else ""
            if mail and name and "@" in mail:
                addr_map.setdefault(mail, name)
    wb.close()
    return addr_map

def cc_from_body(body, addr_map):
    for ln in body.splitlines():
        m = re.match(r"\s*참조\s*[:：]\s*(.+)", ln)
        if m:
            addrs = re.findall(r"[\w.+-]+@[\w.-]+\.\w+", m.group(1))
            seen = []
            for a in addrs:
                disp = addr_map.get(a.lower(), a)   # 주소록에 있으면 이름, 없으면 메일 원문
                if disp not in seen:
                    seen.append(disp)
            return ", ".join(seen)
    return ""

def parse_system_sent(body):
    """본문 '보낸 날짜:' (mMDM 시스템 원본 발송시각) → datetime. 실패 시 원문 문자열."""
    for ln in body.splitlines():
        m = re.match(r"\s*보낸 날짜\s*[:：]\s*(.+)", ln)
        if not m:
            continue
        s = m.group(1).strip()
        dm = re.search(
            r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일.*?(오전|오후)?\s*(\d{1,2}):(\d{2})(?::(\d{2}))?",
            s)
        if dm:
            y, mo, d = int(dm.group(1)), int(dm.group(2)), int(dm.group(3))
            ampm = dm.group(4)
            hh, mm = int(dm.group(5)), int(dm.group(6))
            ss = int(dm.group(7) or 0)
            if ampm == "오후" and hh != 12:
                hh += 12
            if ampm == "오전" and hh == 12:
                hh = 0
            try:
                return datetime.datetime(y, mo, d, hh, mm, ss)
            except Exception:
                return s
        return s
    return ""

def gubun(subj):
    # [설비 등록] -> 생성, 그 외(변경/삭제) -> 변경 / 제목에 '외 N' 있으면 BULK
    base = "생성" if re.search(r"\[설비\s*등록\]", subj) else "변경"
    bulk = "BULK" if re.search(r"외\s*\d+", subj) else "단건"
    return f"{base}({bulk})"

# ---------------------------------------------------------------- Outlook 스캔
def force_sync(ns):
    """실행 시 Outlook에 강제 Send/Receive를 걸고 잠깐 대기 → 방금 도착한 메일도 로컬로 내려받게."""
    try:
        sos = ns.SyncObjects
        n = sos.Count
        if n == 0:
            print("동기화 대상 없음(건너뜀)")
            return
        for i in range(1, n + 1):
            try:
                sos.Item(i).Start()
            except Exception:
                pass
        print(f"Outlook 동기화 요청 → {SYNC_WAIT}초 대기...")
        time.sleep(SYNC_WAIT)
    except Exception as e:
        print("동기화 시도 실패(무시하고 진행):", e)

def scan_outlook():
    addr_map = load_addrbook()
    print(f"주소록 {len(addr_map)}건 로드" if addr_map else "주소록 없음 → 메일 원문 사용")
    ns = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
    force_sync(ns)                       # 먼저 최신 메일 동기화
    inbox = ns.GetDefaultFolder(6)
    items = inbox.Items
    items.Sort("[ReceivedTime]", True)
    try:
        items = items.Restrict(
            "@SQL=\"urn:schemas:httpmail:subject\" like '%Workflow ID%'")
    except Exception:
        pass

    records, matched, skipped_before = [], 0, 0
    seen = set()
    for it in items:
        try:
            if getattr(it, "Class", 0) != 43:
                continue
            subj = getattr(it, "Subject", "") or ""
            addr = (smtp_of(it) or "").lower()
            if SENDER_KEY not in addr or "Workflow ID" not in subj:
                continue
            recv = it.ReceivedTime
            recv_d = datetime.date(recv.year, recv.month, recv.day)
            if recv_d < CUTOFF:          # 7월 이전 접수건 제외
                skipped_before += 1
                continue
            matched += 1
            g = gubun(subj)
            body = it.Body or ""
            cc = cc_from_body(body, addr_map)
            sys_sent = parse_system_sent(body)      # mMDM 시스템 원본 발송시각
            eq = parse_equipment_list(body)
            if not eq:
                records.append((recv_d, {"접수일": recv_d, "구분": g,
                                         "등록요청관련": cc, "비고": "설비목록 파싱실패",
                                         "접수 시간": sys_sent}))
                continue
            for r in eq:
                wf = str(r.get("Workflow ID", "") or "")
                eid = r.get("설비ID", "")
                key = (wf, eid)
                if key in seen and key != ("", ""):
                    continue
                seen.add(key)
                records.append((recv_d, {
                    "접수일": recv_d,
                    "구분": g,
                    "Workflow ID": wf,
                    "설비ID": eid,
                    "등록요청관련": cc,
                    "접수 시간": sys_sent,
                    "Criticality": r.get("Criticality", ""),
                    "기술식별번호": r.get("기술식별번호", ""),
                }))
        except Exception:
            continue
    return records, matched, skipped_before

# ---------------------------------------------------------------- 쓰기
def write_xlsx(records, out_path):
    col = {name: i for i, name in enumerate(HEADER)}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MOC"
    ws.cell(row=1, column=1, value="▼ mMDM Review 등록/변경 (메일 자동수집)")

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    for cidx, name in enumerate(HEADER, 1):
        c = ws.cell(row=2, column=cidx, value=name)
        c.font = Font(name="맑은 고딕", bold=True, size=10)
        c.fill = hdr_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 3
    for _, d in records:
        ws.cell(row=r, column=1, value=d.get("_no"))   # 생성순 번호(내림차순으로 배치됨)
        for name, val in d.items():
            if name in col:
                cc = ws.cell(row=r, column=col[name] + 1)
                if name == "Workflow ID":
                    sval = str(val).strip()
                    if sval.isdigit():
                        cc.value = int(sval)      # 앞자리 0 제거 → 숫자로 (예: 0000054264 → 54264)
                        cc.number_format = "0"
                    else:
                        cc.value = sval
                elif name == "접수일":
                    cc.value = val
                    cc.number_format = "yyyy-mm-dd"
                elif name == "접수 시간":
                    cc.value = val
                    if isinstance(val, datetime.datetime):
                        cc.number_format = "yyyy-mm-dd hh:mm:ss"
                else:
                    cc.value = val
                cc.font = Font(name="맑은 고딕", size=10)
                cc.border = border
        r += 1

    widths = {"접수일": 12, "완료일(월)": 12, "구분": 11, "Workflow ID": 13, "설비ID": 16,
              "등록요청관련": 22, "비고": 18, "접수 시간": 20, "Criticality": 11, "기술식별번호": 13}
    for name, w in widths.items():
        if name in col:
            ws.column_dimensions[get_column_letter(col[name] + 1)].width = w
    ws.freeze_panes = "A3"
    os.makedirs(OUT_RESULT_DIR, exist_ok=True)
    wb.save(out_path)

# ---------------------------------------------------------------- main
def main():
    records, matched, skipped = scan_outlook()
    def asc_key(x):                             # 1차 접수일, 2차 접수시간 (오름차순 기준)
        dkey = x[0] if isinstance(x[0], datetime.date) else datetime.date.min
        t = x[1].get("접수 시간")
        tkey = t if isinstance(t, datetime.datetime) else datetime.datetime.min
        return (dkey, tkey)
    records.sort(key=asc_key)                    # 접수일 → 접수시간 오름차순
    for i, (_, d) in enumerate(records, 1):     # no.는 생성순(오래된 게 1)
        d["_no"] = i
    records.reverse()                           # 뒤집기 → 접수일 내림차순, 같은 날은 접수시간 내림차순

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUT_RESULT_DIR, f"mMDM_MOC_메일수집_{ts}.xlsx")
    write_xlsx(records, out_path)

    print(f"CUTOFF: {CUTOFF} 이후만 수집")
    print(f"패턴A 매칭(7월+): {matched}  /  CUTOFF 이전 제외: {skipped}")
    print(f"총 {len(records)}행 저장 완료")
    print(f"저장: {out_path}")

    # 결과 자동 열기: 엑셀 파일 + 결과 폴더(해당 파일 선택) 둘 다
    try:
        os.startfile(out_path)          # 기본앱(Excel)로 파일 열기
        print("→ 엑셀 파일 열기 완료")
    except Exception as e:
        print("엑셀 파일 열기 실패:", e)
    try:
        import subprocess
        subprocess.Popen(f'explorer /select,"{out_path}"')   # 폴더 열고 파일 선택
        print("→ 결과 폴더 열기 완료")
    except Exception as e:
        print("폴더 열기 실패:", e)

if __name__ == "__main__":
    main()
